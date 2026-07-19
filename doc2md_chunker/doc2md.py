# doc2md_chunker/doc2md.py
from __future__ import annotations

import asyncio
import os
import re
import base64
import httpx
from typing import List, Any, Optional
from pathlib import Path
import tempfile
from fastapi import Depends, FastAPI, UploadFile, File, HTTPException, Query
from fastapi.responses import PlainTextResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from loguru import logger
from markitdown import MarkItDown
from chunker import ChunkResponse, ChunkerName, chunk_text, router as chunker_router
from translate import router as translate_router
import shutil
import uuid
from urllib.parse import urlparse
import pymupdf4llm
import docx2md
from pptx2md import convert as pptx2md_convert, ConversionConfig
import csv
from openpyxl import load_workbook
import mammoth
from dotenv import load_dotenv
from security import get_api_key

import utils

ENV_PATH = Path(__file__).resolve().parent.parent / "doc2md_chunker.env"
load_dotenv(ENV_PATH)

HTTPX_TIMEOUT = httpx.Timeout(float(os.getenv("HTTPX_TIMEOUT", "600.0")))

LLM_MODEL_DEFAULT_BASE_URL = os.getenv(
    "LLM_MODEL_DEFAULT_BASE_URL", "http://localhost:11434"
).rstrip("/")

LLM_MODEL_CHAT_BASE_URL = os.getenv(
    "LLM_MODEL_CHAT_BASE_URL", LLM_MODEL_DEFAULT_BASE_URL
).rstrip("/")

LLM_MODEL_TRANSLATE_BASE_URL = os.getenv(
    "LLM_MODEL_TRANSLATE_BASE_URL", LLM_MODEL_DEFAULT_BASE_URL
).rstrip("/")

LLM_MODEL_CHUNK_BASE_URL = os.getenv(
    "LLM_MODEL_CHUNK_BASE_URL", LLM_MODEL_DEFAULT_BASE_URL
).rstrip("/")

LLM_MODEL_CLASSIFIER_BASE_URL = os.getenv(
    "LLM_MODEL_CLASSIFIER_BASE_URL", LLM_MODEL_DEFAULT_BASE_URL
).rstrip("/")

LLM_MODEL_TABLE_BASE_URL = os.getenv(
    "LLM_MODEL_TABLE_BASE_URL", LLM_MODEL_DEFAULT_BASE_URL
).rstrip("/")

LLM_MODEL_OCR_BASE_URL = os.getenv(
    "LLM_MODEL_OCR_BASE_URL", LLM_MODEL_DEFAULT_BASE_URL
).rstrip("/")

LLM_MODEL_PICTURE_BASE_URL = os.getenv(
    "LLM_MODEL_PICTURE_BASE_URL", LLM_MODEL_DEFAULT_BASE_URL
).rstrip("/")
# ---------------------------------------------------------------------------
# Environment / defaults
# ---------------------------------------------------------------------------
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(50 * 1024 * 1024)))
APP_TMP_BASE = Path(os.getenv("APP_TMP_BASE", "/home/riaz/Code/testAPIs/doc2md_chunker/tmp"))
APP_TMP_BASE.mkdir(parents=True, exist_ok=True)
KEEP_TMP = os.getenv("KEEP_TMP", "false").lower() == "true"
KEEP_TMP = True
LOG_DIR = Path(os.getenv("LOG_DIR", "/home/riaz/Code/testAPIs/doc2md_chunker/var/log/app"))
LOG_DIR.mkdir(parents=True, exist_ok=True)
logger.add(LOG_DIR / "app.log", rotation="10 MB", retention=10)
FIXED_TEMPERATURE = float(os.getenv("LLM_OCR_TEMPERATURE", "0.1"))
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*").split(",")

_sync_http_client = httpx.Client(timeout=HTTPX_TIMEOUT)
_async_http_client = httpx.AsyncClient(timeout=HTTPX_TIMEOUT)

IMAGE_EXTS = set(os.getenv("IMAGE_EXTS", ".png,.jpg,.jpeg,.webp,.bmp,.tiff").split(","))
SUPPORTED_EXTS = set(
    os.getenv(
        "SUPPORTED_EXTS",
        ".pdf,.docx,.pptx,.xlsx,.txt,.md,.html,.htm,.csv,.json,.xml",
    ).split(",")
)

# ---------------------------------------------------------------------------
# FastAPI application
# ---------------------------------------------------------------------------
app = FastAPI(
    title="Markdown Converter Service",
    description="""
A comprehensive document conversion and processing API service that converts various document
formats to Markdown, provides image description capabilities, chunking, and translation services.

## Main Features

* **Document Conversion**: Convert PDF, DOCX, PPTX, and other formats to Markdown
* **Universal Convert**: Convert ANY supported file — document, image, or media — in a single call (`/convert/convert_all`)
* **Image Description**: Classify and describe embedded images using vision LLMs (via `/image-description`)
* **Text Chunking**: Split documents into manageable chunks using various strategies
* **Translation**: Translate text and documents between languages
* **Language Detection**: Identify the language of text or documents
* **Media Transcription**: Transcribe audio and video files to text via OpenAI Whisper

## Supported Document Formats

PDF, DOCX, PPTX, XLSX,TXT, MD, HTML, HTM, CSV, JSON, XML

## Supported Image Formats

PNG, JPG, JPEG, WEBP, BMP, TIFF

## Supported Media Formats

**Audio:** MP3, WAV, FLAC, OGG, M4A, AAC, WMA, OPUS, WEBM  
**Video:** MP4, MKV, AVI, MOV, FLV, WMV, TS, 3GP

## Processing Keywords

All processing keywords default to `false` on `/convert/convert` and `/convert/convert_all`:

- `process_ocr` — run OCR on embedded or uploaded text-heavy images
- `process_tables` — describe table and grid-like images
- `process_figures` — describe charts, graphs, diagrams, and figures
- `process_images` — describe general pictures, photos, and illustrations
- `do_voice_to_text` — transcribe audio or video with Whisper on `/convert/convert_all`

Only enabled processors run; disabled image categories are stripped from the Markdown output.
""",
    version="1.0.0",
    contact={"name": "API Support", "email": "support@example.com"},
    openapi_url="/openapi.json",
    docs_url="/docs",
    redoc_url="/redoc",
    license_info={
        "name": "Apache 2.0",
        "url": "https://www.apache.org/licenses/LICENSE-2.0.html",
    },
    dependencies=[Depends(get_api_key)],
    swagger_ui_parameters={
        "persistAuthorization": True,
        "displayRequestDuration": True,
    },
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Compile once
# ---------------------------------------------------------------------------
SPAN_PATTERN = re.compile(r"<span[^>]*>(.*?)</span>", re.DOTALL)
IMAGE_PATTERN = re.compile(r"!\[(?P<alt>[^\]]*)\]\((?P<src>[^)]+)\)")

# ---------------------------------------------------------------------------
# Favicon
# ---------------------------------------------------------------------------
@app.get(
    "/favicon.ico",
    summary="Serve favicon",
    description="Returns a 1×1 transparent PNG as favicon to prevent 404 errors.",
    tags=["Utility"],
    include_in_schema=False,
)
async def favicon():
    empty_png = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f"
        b"\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc`\x00\x00\x00\x02\x00"
        b"\x01\xe2!\xbc3\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    return Response(content=empty_png, media_type="image/png")


from fastapi.responses import RedirectResponse

@app.get("/", include_in_schema=False)
async def root_redirect():
    return RedirectResponse(url="/docs")

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------
def cleanup_tmp_dir(path: Path) -> None:
    if KEEP_TMP:
        return
    try:
        shutil.rmtree(path, ignore_errors=True)
    except Exception as e:
        logger.error(f"Failed to clean up temp dir {path}: {e}")


async def enforce_max_upload_size(
    file: UploadFile, max_bytes: int = MAX_UPLOAD_BYTES
) -> None:
    length_header = file.headers.get("content-length")
    if length_header:
        try:
            length = int(length_header)
            if length > max_bytes:
                raise HTTPException(
                    status_code=413,
                    detail=f"File too large ({length} bytes). Max allowed is {max_bytes} bytes.",
                )
        except ValueError:
            pass


def strip_spans(markdown: str) -> str:
    return SPAN_PATTERN.sub(r"\1", markdown)


def _safe_filename_from_url(url: str, default_prefix: str = "media") -> str:
    parsed = urlparse(url)
    name = os.path.basename(parsed.path)
    if not name or name in ["/", "."]:
        name = default_prefix
    return name


def _guess_ext_from_mime(mime: str) -> str:
    if not mime:
        return ""
    if mime.startswith("image/"):
        return "." + mime.split("/", 1)[1]
    return ""


def _save_unique(path: Path, data: bytes) -> Path:
    base, ext = path.stem, path.suffix
    candidate = path
    counter = 1
    while candidate.exists():
        candidate = path.with_name(f"{base}_{counter}{ext}")
        counter += 1
    candidate.write_bytes(data)
    return candidate


def _download_image(url: str, images_dir: Path) -> str | None:
    try:
        resp = _sync_http_client.get(url)
        resp.raise_for_status()
    except Exception as e:
        logger.warning(f"Failed to download image {url}: {e}")
        return None
    filename = _safe_filename_from_url(url)
    ext = os.path.splitext(filename)[1]
    if not ext:
        ext = _guess_ext_from_mime(resp.headers.get("Content-Type", ""))
    filename += ext
    img_path = _save_unique(images_dir / filename, resp.content)
    return img_path.name


def _decode_data_uri(uri: str, images_dir: Path) -> str | None:
    try:
        header, b64data = uri.split(",", 1)
    except ValueError:
        return None
    mime = header[5:].split(";")[0] if header.startswith("data:") else ""
    ext = _guess_ext_from_mime(mime) or ".bin"
    try:
        raw = base64.b64decode(b64data)
    except Exception:
        return None
    img_path = _save_unique(images_dir / ("embedded" + ext), raw)
    return img_path.name


def rewrite_markdown_images_inplace(markdown: str, images_dir: Path) -> str:
    """Download remote / data-URI images to *images_dir* and rewrite src paths."""
    images_dir.mkdir(exist_ok=True)

    def replace_match(match: re.Match) -> str:
        alt = match.group("alt")
        src = match.group("src").strip()
        local_name = None
        if src.startswith("data:"):
            local_name = _decode_data_uri(src, images_dir)
        elif src.startswith("http://") or src.startswith("https://"):
            local_name = _download_image(src, images_dir)
        if local_name:
            return f"![{alt}](media/{local_name})"
        return match.group(0)

    return IMAGE_PATTERN.sub(replace_match, markdown)

# ---------------------------------------------------------------------------
# Format-specific converters
# ---------------------------------------------------------------------------
def convert_pdf_with_pymupdf(input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    images_dir = output_dir / "media"
    images_dir.mkdir(exist_ok=True)

    import fitz  # PyMuPDF
    doc = fitz.open(str(input_path))
    num_pages = len(doc)
    doc.close()

    parts: list[str] = []
    for page_num in range(num_pages):
        page_md = pymupdf4llm.to_markdown(
            doc=str(input_path),
            pages=[page_num],
            write_images=True,
            image_path=str(images_dir),
            image_format="png",
        )
        parts.append(f"<!-- page: {page_num + 1} -->\n\n{page_md.strip()}")

    md_text = "\n\n".join(parts)
    md_path = output_dir / f"{input_path.stem}.md"
    md_path.write_text(md_text, encoding="utf-8")
    return md_path

def convert_docx_with_mammoth(input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    images_dir = output_dir / "media"
    images_dir.mkdir(exist_ok=True)

    import tempfile as _tempfile
    from docx import Document as _Document
    from docx.oxml.ns import qn as _qn
    from docx.oxml import OxmlElement as _OxmlElement

    _MARKER_RE = re.compile(r"DOCPAGEMARKER(\d+)END")
    _MARKER_PARA_TEXT = "DOCPAGEMARKER{}END"

    def _make_marker_para(n: int):
        p = _OxmlElement("w:p")
        r = _OxmlElement("w:r")
        t = _OxmlElement("w:t")
        t.text = _MARKER_PARA_TEXT.format(n)
        r.append(t)
        p.append(r)
        return p

    def _elem_has_page_break(elem) -> bool:
        for br in elem.iter(_qn("w:br")):
            if br.get(_qn("w:type")) == "page":
                return True
        return False

    # Build a modified DOCX with page-boundary markers injected
    wdoc = _Document(str(input_path))
    body = wdoc.element.body

    # Insert page-1 marker at the very beginning
    body.insert(0, _make_marker_para(1))

    page_counter = 1
    i = 1  # start after the marker we just inserted
    while i < len(body):
        child = body[i]
        if _elem_has_page_break(child):
            page_counter += 1
            body.insert(i + 1, _make_marker_para(page_counter))
            i += 2  # skip the newly inserted marker
        else:
            i += 1

    tmp_docx = Path(_tempfile.mktemp(suffix=".docx"))
    wdoc.save(str(tmp_docx))

    image_counter = [0]

    def _image_handler(image):
        image_counter[0] += 1
        ext = (image.content_type or "image/png").split("/")[-1]
        if ext == "jpeg":
            ext = "jpg"
        filename = f"image_{image_counter[0]}.{ext}"
        with image.open() as img_data:
            (images_dir / filename).write_bytes(img_data.read())
        return {"src": f"media/{filename}"}

    try:
        with open(tmp_docx, "rb") as f:
            result = mammoth.convert_to_markdown(
                f,
                convert_image=mammoth.images.img_element(_image_handler),
            )
    finally:
        tmp_docx.unlink(missing_ok=True)

    def _replace_marker(m: re.Match) -> str:
        n = int(m.group(1))
        return f"\n\n<!-- page: {n} -->\n\n"

    md_text = _MARKER_RE.sub(_replace_marker, result.value)

    md_path = output_dir / f"{input_path.stem}.md"
    md_path.write_text(md_text, encoding="utf-8")
    return md_path

def convert_docx_with_docx2md(input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    md_text = docx2md.do_convert(
        str(input_path),
        target_dir=str(output_dir),
        use_md_table=True,
    )
    
    md_path = output_dir / f"{input_path.stem}.md"
    md_path.write_text(md_text, encoding="utf-8")
    return md_path


def rows_to_markdown(rows: list[list]) -> str:
    rows = [["" if v is None else str(v) for v in row] for row in rows]
    if not rows:
        return ""
    header = rows[0]
    body = rows[1:]
    separator = " | ".join(["---"] * len(header))
    body_lines = [" | ".join(r) for r in body]
    return (
        " | ".join(header)
        + "\n" + separator
        + ("\n" + "\n".join(body_lines) if body_lines else "")
    )


def convert_csv_xls_to_md(
    input_path: Path, output_dir: Path, sheet_name: str | None = None
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        with open(input_path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
    elif suffix in {".xls", ".xlsx"}:
        wb = load_workbook(filename=input_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    else:
        raise ValueError(f"Unsupported file type: {suffix}")
    md_path = output_dir / f"{input_path.stem}.md"
    md_path.write_text(rows_to_markdown(rows), encoding="utf-8")
    return md_path


def convert_pptx_with_pptx2md(input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    out_md_path = output_dir / f"{input_path.stem}.md"
    images_dir = output_dir / "media"
    images_dir.mkdir(exist_ok=True)
    cfg = ConversionConfig(
        pptx_path=input_path,
        output_path=out_md_path,
        image_dir=images_dir,
        disable_notes=False,
    )
    pptx2md_convert(cfg)
    text = out_md_path.read_text(encoding="utf-8")

    text = strip_spans(text)

    # Split on horizontal-rule slide separators that pptx2md inserts between slides
    _HR_RE = re.compile(r"(?:^|\n)[ \t]*---[ \t]*(?:\n|$)")
    slides = _HR_RE.split(text)

    parts: list[str] = []
    for slide_num, slide_content in enumerate(slides, 1):
        content = slide_content.strip()
        if content:
            parts.append(f"<!-- page: {slide_num} -->\n\n{content}")

    text = "\n\n---\n\n".join(parts)
    out_md_path.write_text(text, encoding="utf-8")
    return out_md_path

def convert_other_with_markitdown(input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    images_dir = output_dir / "media"
    md = MarkItDown(enable_plugins=False)
    result = md.convert(str(input_path))
    markdown = getattr(result, "text_content", None) or getattr(result, "markdown", "")
    markdown = rewrite_markdown_images_inplace(markdown, images_dir)
    md_path = output_dir / f"{input_path.stem}.md"
    md_path.write_text(markdown, encoding="utf-8")
    return md_path


def convert_any_document(input_path: str | Path, output_dir: str | Path = ".") -> Path:
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    suffix = input_path.suffix.lower()
    if suffix == ".pdf":
        return convert_pdf_with_pymupdf(input_path, output_dir)
    if suffix in (".docx", ".doc"):
        return convert_docx_with_mammoth(input_path, output_dir)
    if suffix in (".pptx", ".ppt"):
        return convert_pptx_with_pptx2md(input_path, output_dir)
    if suffix in (".csv", ".xls", ".xlsx"):
        return convert_csv_xls_to_md(input_path, output_dir)
    return convert_other_with_markitdown(input_path, output_dir)


def any_processing_enabled(
    *,
    process_ocr: bool = False,
    process_tables: bool = False,
    process_figures: bool = False,
    process_images: bool = False,
    do_voice_to_text: bool = False,
) -> bool:
    return any(
        (
            process_ocr,
            process_tables,
            process_figures,
            process_images,
            do_voice_to_text,
        )
    )


# ---------------------------------------------------------------------------
# Image description — delegates to utils
# ---------------------------------------------------------------------------
def _describe_image_at_path(
    img_path: Path,
    *,
    temperature: float,
    model_classifier: str,
    model_ocr: str,
    model_table: str,
    model_figure: str,
    model_picture: str,
    process_ocr: bool = False,
    process_tables: bool = False,
    process_figures: bool = False,
    process_images: bool = False,
) -> str:
    try:
        b64 = utils._encode_image_b64(img_path)
        category = utils.classify_image(b64, model_classifier, temperature)
        logger.info(f"Image {img_path.name} classified as {category.value}")

        if category == utils.ImageCategory.text:
            if not process_ocr:
                return ""
            return utils._describe_text_ocr(b64, model_ocr, temperature)
        if category == utils.ImageCategory.table:
            if not process_tables:
                return ""
            return utils._describe_table(b64, model_table, temperature)
        if category == utils.ImageCategory.figure:
            if not process_figures:
                return ""
            return utils._describe_figure(b64, model_figure, temperature)
        if not process_images:
            return ""
        return utils._describe_picture(b64, model_picture, temperature)
    except Exception as exc:
        logger.exception(f"Image description failed for {img_path}: {exc}")
        return ""


def process_markdown_images(
    markdown: str,
    *,
    base_dir: Path | None = None,   
    process_ocr: bool = False,
    process_tables: bool = False,
    process_figures: bool = False,
    process_images: bool = False,
    temperature: float,
    model_classifier: str,
    model_ocr: str,
    model_table: str,
    model_figure: str,
    model_picture: str,
) -> str:

    def _repl(match: re.Match) -> str:
        alt_text = (match.group("alt") or "").strip()
        img_path_str = match.group("src").strip()

        if not any_processing_enabled(
            process_ocr=process_ocr,
            process_tables=process_tables,
            process_figures=process_figures,
            process_images=process_images,
        ):
            return ""

        img_path = Path(img_path_str)
        # ↓ FIX: resolve relative paths against the document's output directory
        if not img_path.is_absolute() and base_dir is not None:
            img_path = (base_dir / img_path).resolve()

        if not img_path.exists():
            logger.warning(f"Image not found: {img_path} — skipping description")
            return ""

        description = _describe_image_at_path(
            img_path,
            temperature=temperature,
            model_classifier=model_classifier,
            model_ocr=model_ocr,
            model_table=model_table,
            model_figure=model_figure,
            model_picture=model_picture,
            process_ocr=process_ocr,
            process_tables=process_tables,
            process_figures=process_figures,
            process_images=process_images,
        )
        if not description:
            return ""

        label = alt_text if alt_text else "Image"
        return f"**[{label}]** {description}"

    return IMAGE_PATTERN.sub(_repl, markdown)


# ---------------------------------------------------------------------------
# /convert/convert
# ---------------------------------------------------------------------------
@app.post(
    "/convert/convert",
    response_class=PlainTextResponse,
    summary="Convert document to Markdown",
    description="""
Upload a document and convert it to Markdown.

## Processing Keywords

All processing keywords default to `false`.

- `process_ocr` — process embedded text-heavy images with the OCR model.
- `process_tables` — process embedded table and grid images with the table model.
- `process_figures` — process embedded figures, charts, and diagrams with the figure model.
- `process_images` — process embedded photographs, illustrations, and general picture-like images.
- `do_voice_to_text` — accepted for API consistency but ignored on this endpoint because `/convert/convert` only accepts document uploads.

Only enabled image categories are described inline. Disabled image categories are stripped from the output.
Image links are never included in the output.

## Supported Formats

PDF, DOCX, PPTX, XLSX, TXT, MD, HTML, HTM, CSV, JSON, XML

## Error Codes

- **413** File too large
- **400** Unsupported file format
- **500** Conversion or LLM error (check service logs)
""",
    tags=["Conversion"],
    response_description="Markdown text (no image links)",
)
async def convert_file(
    file: UploadFile = File(..., description="Document file to convert to Markdown."),
    process_ocr: bool = Query(
        False,
        description="When true, OCR/transcribe embedded text-heavy images. Disabled by default.",
    ),
    process_tables: bool = Query(
        False,
        description="When true, describe embedded tables and structured grid images. Disabled by default.",
    ),
    process_figures: bool = Query(
        False,
        description="When true, describe embedded figures, charts, and diagrams. Disabled by default.",
    ),
    process_images: bool = Query(
        False,
        description="When true, describe embedded photographs, illustrations, and general pictures. Disabled by default.",
    ),
    do_voice_to_text: bool = Query(
        False,
        description="Ignored on `/convert/convert` because this endpoint only processes documents.",
    ),
    temperature: float = Query(
        FIXED_TEMPERATURE,
        ge=0.0,
        le=1.0,
        description="Sampling temperature for enabled vision LLM calls (0 = deterministic).",
    ),
    model_classifier: str = Query(
        utils.MODEL_CLASSIFIER,
        description="Vision model used to classify each image into Picture / Text / Table / Figure.",
    ),
    model_picture: str = Query(
        utils.MODEL_PICTURE,
        description="Vision model used when `process_images=true` and the image is classified as Picture.",
    ),
    model_ocr: str = Query(
        utils.MODEL_OCR,
        description="Vision model used when `process_ocr=true` and the image is classified as Text/OCR.",
    ),
    model_table: str = Query(
        utils.MODEL_TABLE,
        description="Vision model used when `process_tables=true` and the image is classified as Table.",
    ),
    model_figure: str = Query(
        utils.MODEL_FIGURE,
        description="Vision model used when `process_figures=true` and the image is classified as Figure.",
    ),
):
    await enforce_max_upload_size(file)
    suffix = Path(file.filename).suffix.lower()
    if suffix not in SUPPORTED_EXTS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported extension {suffix!r}. Allowed: {sorted(SUPPORTED_EXTS)}",
        )

    logger.info(
        f"convert called: filename={file.filename}, "
        f"process_ocr={process_ocr}, process_tables={process_tables}, "
        f"process_figures={process_figures}, process_images={process_images}, "
        f"do_voice_to_text={do_voice_to_text}, temperature={temperature}, "
        f"model_classifier={model_classifier}"
    )

    tmp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp_path = Path(tmp.name)
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                tmp.write(chunk)

        req_id = uuid.uuid4().hex
        out_dir = APP_TMP_BASE / req_id
        out_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"file: {tmp_path} | output dir: {out_dir}")

        try:
            md_path = convert_any_document(tmp_path, out_dir)
            raw_text = md_path.read_text(encoding="utf-8").strip()

            result = process_markdown_images(
                raw_text,
                base_dir=out_dir,
                process_ocr=process_ocr,
                process_tables=process_tables,
                process_figures=process_figures,
                process_images=process_images,
                temperature=temperature,
                model_classifier=model_classifier,
                model_ocr=model_ocr,
                model_table=model_table,
                model_figure=model_figure,
                model_picture=model_picture,
            )

            logger.info(
                f"Converted {file.filename} → {len(result)} chars"
                f" (ocr={process_ocr}, tables={process_tables}, figures={process_figures}, images={process_images})"
            )

        finally:
            cleanup_tmp_dir(out_dir)

        return result

    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception as e:
                logger.error(f"Failed to clean up temp file {tmp_path}: {e}")

# ---------------------------------------------------------------------------
# /convert/convert_all  (NEW)
# Combines /convert/convert  (documents + images)
#      and /media_to_text     (audio + video via Whisper)
# into a single universal endpoint.
# ---------------------------------------------------------------------------
@app.post(
    "/convert/convert_all",
    response_class=PlainTextResponse,
    summary="Universal file-to-text converter (documents, images, audio, video)",
    description="""
A single endpoint that accepts any supported file type and routes it to the
appropriate conversion pipeline automatically.

| File category | Accepted extensions | Pipeline used |
|---------------|---------------------|---------------|
| **Document** | PDF, DOCX, PPTX, XLSX, TXT, MD, HTML, HTM, CSV, JSON, XML | Document → Markdown |
| **Image** | PNG, JPG, JPEG, WEBP, BMP, TIFF | Category-aware vision processing |
| **Audio** | MP3, WAV, FLAC, OGG, M4A, AAC, WMA, OPUS, WEBM | Whisper speech-to-text when `do_voice_to_text=true` |
| **Video** | MP4, MKV, AVI, MOV, FLV, WMV, TS, 3GP | Whisper speech-to-text when `do_voice_to_text=true` |

## Processing Keywords

All processing keywords default to `false`.

- `process_ocr` — enable OCR processing for text-heavy images.
- `process_tables` — enable table extraction or description for table-like images.
- `process_figures` — enable figure, chart, and diagram description.
- `process_images` — enable picture, photo, and illustration description.
- `do_voice_to_text` — enable audio and video transcription via Whisper.

## Output Format

- **Document / Image** → clean Markdown text with no raw image links.
- **Audio / Video** → plain-text transcription wrapped in HTML comment anchors.

## Error Codes

- **413** File too large (default limit: 100 MB)
- **400** Unsupported file extension or media transcription disabled
- **500** Conversion, LLM, or Whisper error — check service logs
""",
    tags=["Conversion"],
    response_description=(
        "Markdown text for documents/images, or plain-text transcription for audio/video."
    ),
)
async def convert_all(
    file: UploadFile = File(
        ...,
        description=(
            "Any supported file: document (PDF, DOCX, PPTX, …), "
            "image (PNG, JPG, …), audio (MP3, WAV, …), or video (MP4, MKV, …)."
        ),
    ),
    process_ocr: bool = Query(
        False,
        description="Enable OCR or transcription for text-heavy images. Disabled by default.",
    ),
    process_tables: bool = Query(
        False,
        description="Enable table-image processing. Disabled by default.",
    ),
    process_figures: bool = Query(
        False,
        description="Enable figure, chart, and diagram processing. Disabled by default.",
    ),
    process_images: bool = Query(
        False,
        description="Enable general picture, photograph, and illustration processing. Disabled by default.",
    ),
    do_voice_to_text: bool = Query(
        False,
        description="When true, audio/video files are transcribed with Whisper. Disabled by default.",
    ),
    temperature: float = Query(
        FIXED_TEMPERATURE,
        ge=0.0,
        le=1.0,
        description="Sampling temperature for enabled vision LLM calls (documents/images). 0 = deterministic.",
    ),
    model_classifier: str = Query(
        utils.MODEL_CLASSIFIER,
        description="Vision model for image classification (Picture / Text / Table / Figure).",
    ),
    model_picture: str = Query(
        utils.MODEL_PICTURE,
        description="Vision model for Picture-category images when `process_images=true`.",
    ),
    model_ocr: str = Query(
        utils.MODEL_OCR,
        description="Vision model for Text/OCR-category images when `process_ocr=true`.",
    ),
    model_table: str = Query(
        utils.MODEL_TABLE,
        description="Vision model for Table-category images when `process_tables=true`.",
    ),
    model_figure: str = Query(
        utils.MODEL_FIGURE,
        description="Vision model for Figure/chart-category images when `process_figures=true`.",
    ),
    whisper_model_size: str = Query(
        utils.WHISPER_MODEL_SIZE,
        description=(
            "Whisper model size for audio/video transcription: "
            "`tiny`, `base`, `small`, `medium`, `large`, `large-v2`, `large-v3`. "
            "Ignored unless `do_voice_to_text=true`."
        ),
    ),
    whisper_language: Optional[str] = Query(
        None,
        description=(
            "BCP-47 language code to force for Whisper, e.g. `en`, `fr`, `ja`. "
            "Ignored unless `do_voice_to_text=true`."
        ),
    ),
    whisper_temperature: float = Query(
        0.0,
        ge=0.0,
        le=1.0,
        description=(
            "Whisper decoding temperature. `0.0` = greedy/deterministic (recommended). "
            "Ignored unless `do_voice_to_text=true`."
        ),
    ),
) -> str:
    await enforce_max_upload_size(file)

    suffix = Path(file.filename or "").suffix.lower()

    all_supported = SUPPORTED_EXTS | IMAGE_EXTS | utils.MEDIA_EXTS
    if suffix not in all_supported:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Unsupported extension {suffix!r}. "
                f"Documents: {sorted(SUPPORTED_EXTS)}. "
                f"Images: {sorted(IMAGE_EXTS)}. "
                f"Audio: {sorted(utils.AUDIO_EXTS)}. "
                f"Video: {sorted(utils.VIDEO_EXTS)}."
            ),
        )

    logger.info(
        f"convert_all called: filename={file.filename!r}, suffix={suffix!r}, "
        f"process_ocr={process_ocr}, process_tables={process_tables}, "
        f"process_figures={process_figures}, process_images={process_images}, "
        f"do_voice_to_text={do_voice_to_text}, whisper_model={whisper_model_size}"
    )

    if suffix in utils.MEDIA_EXTS:
        if not do_voice_to_text:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Received an audio/video file but `do_voice_to_text=false`. "
                    "Set `do_voice_to_text=true` to transcribe media files."
                ),
            )

        tmp_path: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp_path = Path(tmp.name)
                while True:
                    chunk = await file.read(1024 * 1024)
                    if not chunk:
                        break
                    tmp.write(chunk)

            result: dict = await asyncio.to_thread(
                utils._transcribe_sync,
                tmp_path,
                whisper_model_size,
                whisper_language,
                whisper_temperature,
            )

            detected_lang: str = result.get("language", "unknown")
            text: str = (result.get("text") or "").strip()
            segments: list = result.get("segments") or []
            duration_sec = float(segments[-1]["end"]) if segments else 0.0
            duration_str = utils._fmt_duration(duration_sec)
            slug = utils._slugify(Path(file.filename or "media").stem)

            logger.info(
                f"convert_all (media) done — language={detected_lang} "
                f"duration={duration_str} chars={len(text)}"
            )

            return (
                f'<!-- media_start slug={slug} language={detected_lang} duration={duration_str} -->'
                f"{text}"
                f"<!-- media_end -->"
            )

        except HTTPException:
            raise
        except Exception as exc:
            logger.exception(f"convert_all transcription failed: {exc}")
            raise HTTPException(status_code=500, detail=f"Transcription failed: {exc}") from exc
        finally:
            if tmp_path and tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception as exc:
                    logger.error(f"Failed to clean up temp file {tmp_path}: {exc}")

    if suffix in IMAGE_EXTS:
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp_path = Path(tmp.name)
                while True:
                    chunk = await file.read(1024 * 1024)
                    if not chunk:
                        break
                    tmp.write(chunk)

            description = _describe_image_at_path(
                tmp_path,
                temperature=temperature,
                model_classifier=model_classifier,
                model_ocr=model_ocr,
                model_table=model_table,
                model_figure=model_figure,
                model_picture=model_picture,
                process_ocr=process_ocr,
                process_tables=process_tables,
                process_figures=process_figures,
                process_images=process_images,
            )

            logger.info(f"convert_all (image) done — chars={len(description)}")
            return description or ""

        except HTTPException:
            raise
        except Exception as exc:
            logger.exception(f"convert_all image description failed: {exc}")
            raise HTTPException(status_code=500, detail=f"Image description failed: {exc}") from exc
        finally:
            if tmp_path and tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception as exc:
                    logger.error(f"Failed to clean up temp file {tmp_path}: {exc}")

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp_path = Path(tmp.name)
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                tmp.write(chunk)

        req_id = uuid.uuid4().hex
        out_dir = APP_TMP_BASE / req_id
        out_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"convert_all (document): input={tmp_path} out_dir={out_dir}")

        try:
            md_path = convert_any_document(tmp_path, out_dir)
            raw_text = md_path.read_text(encoding="utf-8").strip()

            result_md = process_markdown_images(
                raw_text,
                base_dir=out_dir,
                process_ocr=process_ocr,
                process_tables=process_tables,
                process_figures=process_figures,
                process_images=process_images,
                temperature=temperature,
                model_classifier=model_classifier,
                model_ocr=model_ocr,
                model_table=model_table,
                model_figure=model_figure,
                model_picture=model_picture,
            )

        finally:
            cleanup_tmp_dir(out_dir)

        logger.info(
            f"convert_all (document) {file.filename} → {len(result_md)} chars"
            f" (ocr={process_ocr}, tables={process_tables}, figures={process_figures}, images={process_images})"
        )

        return result_md

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"convert_all document conversion failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Conversion failed: {exc}") from exc
    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception as e:
                logger.error(f"Failed to clean up temp file {tmp_path}: {e}")

# ---------------------------------------------------------------------------
# /convert/doc2md_chunker
# ---------------------------------------------------------------------------
@app.post(
    "/convert/doc2md_chunker",
    response_model=ChunkResponse,
    summary="Convert document to Markdown and chunk",
    description="""
Upload a document, convert it to Markdown, and split it into chunks in a single operation.

This endpoint combines document conversion with text chunking:

## Chunking Strategies

| Method      | Description                                          |
|-------------|------------------------------------------------------|
| `token`     | Split by token count; respects model context windows |
| `sentence`  | Maintain sentence boundaries                         |
| `recursive` | Hierarchical split with custom separators (default)  |
| `semantic`  | Content-aware split by semantic similarity           |
| `late`      | Late-interaction semantic chunking                   |
| `code`      | Language-aware code splitting                        |

## Response Format

```json
{
  "chunks": [
    {"text": "...", "index": 0, "token_count": 150, "description": "Chunk 0, 150 tokens"},
    ...
  ]
}
```

## Error Codes

- **413** File too large
- **400** Unsupported format or invalid chunking parameters
- **500** Conversion or chunking error
""",
    tags=["Conversion"],
    response_description="Structured array of text chunks with metadata",
)
async def doc2md_chunker(
    file: UploadFile = File(..., description="Any supported document to convert and chunk."),
    method: ChunkerName = Query(
        "recursive",
        description="Chunking strategy: token | sentence | recursive | semantic | late | code.",
    ),
    chunk_size: int | None = Query(None, description="Target chunk size in tokens/characters."),
    chunk_overlap: int | None = Query(None, description="Overlap between consecutive chunks."),
    max_tokens: int | None = Query(None, description="Maximum tokens per chunk."),
    separators: List[str] | None = Query(
        None, description="Custom separators for recursive chunking."
    ),
    similarity_threshold: float | None = Query(
        None, description="Similarity threshold for semantic/late chunkers."
    ),
    min_chunk_size: int | None = Query(
        None, description="Minimum chunk size for semantic methods."
    ),
    max_chunk_size: int | None = Query(
        None, description="Maximum chunk size for semantic methods."
    ),
    language: str | None = Query(None, description="Language for code chunking (e.g. \'python\')."),
    tokenizer: str | None = Query(None, description="Tokenizer identifier."),
):
    await enforce_max_upload_size(file)
    suffix = Path(file.filename).suffix.lower()
    if suffix not in SUPPORTED_EXTS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported extension {suffix!r}. Allowed: {sorted(SUPPORTED_EXTS)}",
        )

    tmp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp_path = Path(tmp.name)
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                tmp.write(chunk)

        req_id = uuid.uuid4().hex
        out_dir = APP_TMP_BASE / req_id
        out_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"doc2md_chunker: input={tmp_path} out_dir={out_dir}")

        try:
            md_path = convert_any_document(tmp_path, out_dir)
            text = md_path.read_text(encoding="utf-8").strip()
        finally:
            cleanup_tmp_dir(out_dir)

        raw_kwargs: dict[str, Any] = {
            "chunk_size": chunk_size,
            "chunk_overlap": chunk_overlap,
            "max_tokens": max_tokens,
            "separators": separators,
            "similarity_threshold": similarity_threshold,
            "min_chunk_size": min_chunk_size,
            "max_chunk_size": max_chunk_size,
            "language": language,
            "tokenizer": tokenizer,
        }
        kwargs = {k: v for k, v in raw_kwargs.items() if v is not None}

        try:
            chunks = chunk_text(text, method=method, **kwargs)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except TypeError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Chunker parameter error for method {method!r}: {e}",
            )

        payload = []
        for idx, ch in enumerate(chunks):
            ch_text = getattr(ch, "text", str(ch))
            ch_tokens = getattr(ch, "token_count", None)
            ch_index = getattr(ch, "index", idx)
            desc_parts = [f"Chunk {ch_index}"]
            desc_parts.append(
                f"{ch_tokens} tokens" if ch_tokens is not None else f"{len(ch_text)} characters"
            )

            payload.append(
                {
                    "text": ch_text,
                    "index": ch_index,
                    "token_count": ch_tokens,
                    "description": ", ".join(desc_parts),
                }
            )

        return ChunkResponse(chunks=payload)

    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Include routers — utils provides /health, /features, /image-description,
#                   /media_to_text, /list_llm_models, /running_llm_models
# ---------------------------------------------------------------------------
app.include_router(utils.router)
app.include_router(chunker_router, prefix="/chunk")
app.include_router(translate_router, prefix="/translate")
